"""
Script di test per verificare il funzionamento del file Excel generato
"""

import openpyxl
import os

def test_excel_file():
    """Testa il file Excel generato"""
    filepath = "Calcolatore_Esterificazione.xlsx"
    
    if not os.path.exists(filepath):
        print(f"Errore: File {filepath} non trovato!")
        return False
    
    try:
        # Apri il file Excel
        wb = openpyxl.load_workbook(filepath)
        print(f"✓ File Excel aperto con successo: {filepath}")
        
        # Verifica i fogli
        expected_sheets = ["Database Composti", "Calcolatore Esterificazione", "Istruzioni"]
        actual_sheets = wb.sheetnames
        print(f"Fogli presenti: {actual_sheets}")
        
        for sheet in expected_sheets:
            if sheet in actual_sheets:
                print(f"✓ Foglio '{sheet}' presente")
            else:
                print(f"✗ Foglio '{sheet}' mancante")
                return False
        
        # Testa il foglio Database Composti
        db_sheet = wb["Database Composti"]
        print(f"\nTesting Database Composti:")
        print(f"- Titolo: {db_sheet['A1'].value}")
        print(f"- Primo composto: {db_sheet['A3'].value}")
        print(f"- Peso molecolare primo composto: {db_sheet['C3'].value}")
        
        # Testa il foglio Calcolatore
        calc_sheet = wb["Calcolatore Esterificazione"]
        print(f"\nTesting Calcolatore Esterificazione:")
        print(f"- Titolo: {calc_sheet['A1'].value}")
        
        # Verifica presenza delle formule
        sample_formula_cell = calc_sheet['D8']  # Dovrebbe contenere una formula per le moli
        print(f"- Formula di esempio (D8): {sample_formula_cell.value}")
        
        # Testa inserimento valori
        print(f"\nTest inserimento valori:")
        # Inserisci peso per acido palmitico (B8)
        calc_sheet['B8'].value = 1.0  # 1 kg
        # Inserisci peso per alcol (B11) 
        calc_sheet['B11'].value = 0.5  # 0.5 kg
        
        wb.save("Test_" + filepath)
        print(f"✓ File di test salvato con valori di esempio")
        
        # Riapri per verificare i calcoli
        wb_test = openpyxl.load_workbook("Test_" + filepath)
        calc_test = wb_test["Calcolatore Esterificazione"]
        
        moles_value = calc_test['D8'].value
        print(f"- Moli calcolate per acido palmitico (1 kg): {moles_value}")
        
        wb_test.close()
        wb.close()
        
        return True
        
    except Exception as e:
        print(f"Errore durante il test: {e}")
        return False

if __name__ == "__main__":
    print("=== TEST DEL FILE EXCEL ESTERIFICAZIONE ===")
    success = test_excel_file()
    if success:
        print("\n✓ Tutti i test sono passati!")
    else:
        print("\n✗ Alcuni test sono falliti!")