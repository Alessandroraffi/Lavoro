"""
Generatore di file Excel per il calcolo automatico dell'acqua prodotta 
in reazioni di esterificazione.

Questo script crea un file Excel con:
- Database dei pesi molecolari dei composti
- Tabelle per l'inserimento dei dati in kg
- Calcoli automatici delle moli
- Calcolo della produzione di acqua
- Controlli di bilancio di massa
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os

class EsterificationCalculator:
    def __init__(self):
        self.wb = Workbook()
        
        # Database dei pesi molecolari (g/mol)
        self.molecular_weights = {
            # Acidi grassi saturi
            'Acido laurico (C12:0)': 200.32,
            'Acido miristico (C14:0)': 228.37,
            'Acido palmitico (C16:0)': 256.42,
            'Acido stearico (C18:0)': 284.48,
            'Acido arachidico (C20:0)': 312.53,
            
            # Acidi grassi insaturi
            'Acido oleico (C18:1)': 282.46,
            'Acido linoleico (C18:2)': 280.45,
            'Acido linolenico (C18:3)': 278.43,
            'Acido arachididonico (C20:4)': 304.47,
            
            # Altri composti
            'Alcol due etil esilico (2-etilesanolo)': 130.23,
            'Glicerina': 92.09,
            'Acido adipico': 146.14,
            'Acqua': 18.015
        }
        
        # Stili per le celle
        self.header_font = Font(bold=True, size=12)
        self.subheader_font = Font(bold=True, size=10)
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.input_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        self.calc_fill = PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def create_header(self, ws, title, row):
        """Crea un'intestazione per una sezione"""
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = self.border
        return row + 1
        
    def create_compound_database(self):
        """Crea il foglio con il database dei composti"""
        ws = self.wb.active
        ws.title = "Database Composti"
        
        # Intestazione
        row = self.create_header(ws, "DATABASE PESI MOLECOLARI COMPOSTI", 1)
        row += 1
        
        # Intestazioni tabella
        headers = ['Composto', 'Formula', 'Peso Molecolare (g/mol)', 'Categoria']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.subheader_font
            cell.fill = self.input_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Dati dei composti
        compounds_data = [
            ('Acido laurico (C12:0)', 'C12H24O2', 200.32, 'Acido grasso saturo'),
            ('Acido miristico (C14:0)', 'C14H28O2', 228.37, 'Acido grasso saturo'),
            ('Acido palmitico (C16:0)', 'C16H32O2', 256.42, 'Acido grasso saturo'),
            ('Acido stearico (C18:0)', 'C18H36O2', 284.48, 'Acido grasso saturo'),
            ('Acido arachidico (C20:0)', 'C20H40O2', 312.53, 'Acido grasso saturo'),
            ('Acido oleico (C18:1)', 'C18H34O2', 282.46, 'Acido grasso insaturo'),
            ('Acido linoleico (C18:2)', 'C18H32O2', 280.45, 'Acido grasso insaturo'),
            ('Acido linolenico (C18:3)', 'C18H30O2', 278.43, 'Acido grasso insaturo'),
            ('Acido arachididonico (C20:4)', 'C20H32O2', 304.47, 'Acido grasso insaturo'),
            ('Alcol due etil esilico', 'C8H18O', 130.23, 'Alcol'),
            ('Glicerina', 'C3H8O3', 92.09, 'Polialcol'),
            ('Acido adipico', 'C6H10O4', 146.14, 'Acido dicarbossilico'),
            ('Acqua', 'H2O', 18.015, 'Prodotto di reazione')
        ]
        
        for compound, formula, mw, category in compounds_data:
            for col, value in enumerate([compound, formula, mw, category], 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col == 3:  # Peso molecolare
                    cell.alignment = Alignment(horizontal='center')
            row += 1
        
        # Ridimensiona le colonne
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25
            
    def create_calculation_sheet(self):
        """Crea il foglio principale per i calcoli"""
        ws = self.wb.create_sheet("Calcolatore Esterificazione")
        
        row = 1
        row = self.create_header(ws, "CALCOLATORE ESTERIFICAZIONE - PRODUZIONE ACQUA", row)
        row += 1
        
        # Sezione reagenti
        row = self.create_header(ws, "REAGENTI", row)
        row += 1
        
        # Intestazioni per i reagenti
        headers = ['Composto', 'Peso (kg)', 'PM (g/mol)', 'Moli (mol)', 'Equivalenti']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.subheader_font
            cell.fill = self.input_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        
        # Lista dei reagenti principali
        reagents = [
            'Acido palmitico (C16:0)',
            'Acido stearico (C18:0)', 
            'Acido oleico (C18:1)',
            'Alcol due etil esilico (2-etilesanolo)',
            'Glicerina',
            'Acido adipico'
        ]
        
        reagent_rows = {}
        for reagent in reagents:
            reagent_rows[reagent] = row
            # Nome composto
            cell = ws.cell(row=row, column=1, value=reagent)
            cell.border = self.border
            
            # Peso (input dell'utente)
            cell = ws.cell(row=row, column=2)
            cell.fill = self.input_fill
            cell.border = self.border
            cell.value = 0  # Valore predefinito
            
            # Peso molecolare (lookup dal database)
            cell = ws.cell(row=row, column=3, value=self.molecular_weights[reagent])
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
            
            # Moli (calcolate automaticamente)
            cell = ws.cell(row=row, column=4)
            cell.value = f'=B{row}*1000/C{row}'  # kg -> g, poi /PM
            cell.border = self.border
            cell.fill = self.calc_fill
            cell.alignment = Alignment(horizontal='center')
            
            # Equivalenti (per ora uguali alle moli)
            cell = ws.cell(row=row, column=5)
            cell.value = f'=D{row}'
            cell.border = self.border
            cell.fill = self.calc_fill
            cell.alignment = Alignment(horizontal='center')
            
            row += 1
        
        row += 1
        
        # Sezione calcoli stechiometrici
        row = self.create_header(ws, "CALCOLI STECHIOMETRICI", row)
        row += 1
        
        # Tipo di reazione
        ws.cell(row=row, column=1, value="Tipo di reazione:")
        ws.cell(row=row, column=2, value="Esterificazione")
        row += 1
        
        # Equazione generale: Acido + Alcol → Estere + Acqua
        ws.cell(row=row, column=1, value="Equazione:")
        ws.cell(row=row, column=2, value="R-COOH + R'-OH → R-COO-R' + H2O")
        row += 2
        
        # Calcolo acqua prodotta
        ws.cell(row=row, column=1, value="Moli di acidi totali:")
        # Somma delle moli degli acidi grassi
        acid_rows = [reagent_rows[r] for r in reagents if 'Acido' in r and 'adipico' not in r]
        acids_formula = '+'.join([f'D{r}' for r in acid_rows])
        cell = ws.cell(row=row, column=2)
        cell.value = f'={acids_formula}'
        cell.fill = self.calc_fill
        cell.border = self.border

        row += 1
        
        ws.cell(row=row, column=1, value="Moli di alcoli totali:")
        # Somma delle moli degli alcoli
        alcohol_rows = [reagent_rows[r] for r in reagents if 'Alcol' in r or 'Glicerina' in r]
        alcohols_formula = '+'.join([f'D{r}' for r in alcohol_rows])  
        cell = ws.cell(row=row, column=2)
        cell.value = f'={alcohols_formula}'
        cell.fill = self.calc_fill
        cell.border = self.border
        
        row += 1
        
        ws.cell(row=row, column=1, value="Reagente limitante (moli):")
        cell = ws.cell(row=row, column=2)
        cell.value = f'=MIN(B{row-2}, B{row-1})'
        cell.fill = self.calc_fill
        cell.border = self.border
        
        limiting_reagent_row = row
        row += 2
        
        # Produzione di acqua
        row = self.create_header(ws, "PRODUZIONE ACQUA", row)
        row += 1
        
        ws.cell(row=row, column=1, value="Moli di acqua prodotta:")
        cell = ws.cell(row=row, column=2)
        cell.value = f'=B{limiting_reagent_row}'
        cell.fill = self.calc_fill
        cell.border = self.border
        
        water_moles_row = row
        row += 1
        
        ws.cell(row=row, column=1, value="Massa acqua prodotta (g):")
        cell = ws.cell(row=row, column=2)
        cell.value = f'=B{water_moles_row}*{self.molecular_weights["Acqua"]}'
        cell.fill = self.calc_fill
        cell.border = self.border
        
        row += 1
        
        ws.cell(row=row, column=1, value="Massa acqua prodotta (kg):")
        cell = ws.cell(row=row, column=2)
        cell.value = f'=B{row-1}/1000'
        cell.fill = self.calc_fill
        cell.border = self.border
        
        row += 2
        
        # Controllo bilancio di massa
        row = self.create_header(ws, "CONTROLLO BILANCIO DI MASSA", row)
        row += 1
        
        ws.cell(row=row, column=1, value="Massa reagenti totale (kg):")
        reagents_sum = '+'.join([f'B{reagent_rows[r]}' for r in reagents])
        cell = ws.cell(row=row, column=2)
        cell.value = f'={reagents_sum}'
        cell.fill = self.calc_fill
        cell.border = self.border
        
        total_reagents_row = row
        row += 1
        
        ws.cell(row=row, column=1, value="Massa prodotti stimata (kg):")
        ws.cell(row=row, column=2, value="=(massa reagenti - massa acqua prodotta)")
        row += 1
        
        ws.cell(row=row, column=1, value="Differenza bilancio (%):")
        cell = ws.cell(row=row, column=2)
        cell.value = f'=ABS(B{total_reagents_row}-B{water_moles_row+2})/B{total_reagents_row}*100'
        cell.fill = self.calc_fill
        cell.border = self.border
        
        # Ridimensiona le colonne
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
            
    def create_instructions_sheet(self):
        """Crea un foglio con le istruzioni d'uso"""
        ws = self.wb.create_sheet("Istruzioni")
        
        row = 1
        row = self.create_header(ws, "ISTRUZIONI D'USO", row)
        row += 2
        
        instructions = [
            "1. INSERIMENTO DATI:",
            "   - Vai al foglio 'Calcolatore Esterificazione'",
            "   - Inserisci i pesi dei reagenti in kg nelle celle grigie della colonna B",
            "   - I calcoli si aggiorneranno automaticamente",
            "",
            "2. INTERPRETAZIONE RISULTATI:",
            "   - Le moli vengono calcolate automaticamente dai pesi inseriti",
            "   - Il reagente limitante determina la quantità di acqua prodotta",
            "   - Il bilancio di massa aiuta a verificare la coerenza dei dati",
            "",
            "3. COMPOSTI SUPPORTATI:",
            "   - Acidi grassi saturi: C12:0, C14:0, C16:0, C18:0, C20:0",
            "   - Acidi grassi insaturi: C18:1, C18:2, C18:3, C20:4",  
            "   - Alcol due etil esilico (2-etilesanolo)",
            "   - Glicerina",
            "   - Acido adipico",
            "",
            "4. REAZIONE DI ESTERIFICAZIONE:",
            "   - Equazione generale: R-COOH + R'-OH → R-COO-R' + H2O",
            "   - Ogni mole di acido reagisce con una mole di alcol",
            "   - Si produce una mole di acqua per ogni mole di legame estereo formato",
            "",
            "5. CONTROLLI DI QUALITÀ:",
            "   - Il bilancio di massa dovrebbe essere vicino al 100%",
            "   - Differenze significative indicano errori nei dati inseriti",
            "   - Verificare che i reagenti siano in proporzioni realistiche"
        ]
        
        for instruction in instructions:
            ws.cell(row=row, column=1, value=instruction)
            if instruction.startswith(("1.", "2.", "3.", "4.", "5.")):
                ws.cell(row=row, column=1).font = self.subheader_font
            row += 1
        
        # Ridimensiona la colonna
        ws.column_dimensions['A'].width = 80
        
    def generate_excel_file(self, filename="Calcolatore_Esterificazione.xlsx"):
        """Genera il file Excel completo"""
        # Crea i fogli
        self.create_compound_database()
        self.create_calculation_sheet()
        self.create_instructions_sheet()
        
        # Salva il file
        filepath = os.path.join(os.getcwd(), filename)
        self.wb.save(filepath)
        print(f"File Excel creato: {filepath}")
        return filepath

def main():
    """Funzione principale"""
    calculator = EsterificationCalculator()
    filepath = calculator.generate_excel_file()
    print("File Excel generato con successo!")
    print(f"Percorso: {filepath}")
    
    # Istruzioni per l'utente
    print("\nISTRUZIONI:")
    print("1. Apri il file Excel generato")
    print("2. Vai al foglio 'Calcolatore Esterificazione'")
    print("3. Inserisci i pesi dei reagenti in kg nelle celle grigie")
    print("4. I calcoli si aggiorneranno automaticamente")
    print("5. Controlla il bilancio di massa per verificare la coerenza")

if __name__ == "__main__":
    main()